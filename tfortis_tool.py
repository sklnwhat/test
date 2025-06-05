import sys
import logging
import json
import ipaddress
import datetime
from typing import Optional
import requests
from requests.auth import HTTPDigestAuth
import pandas as pd
import time
import shutil
import os
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from scapy.all import ARP, Ether, srp

# ——— Настройки ———
SWITCH_USERNAME = 'admin'
SWITCH_PASSWORD = 'top23temr'
CAMERA_USERNAME = 'admin'
CAMERA_PASSWORD = 'a1234567*'

EXCEL_FILE = 'tfortis.xlsx'
BACKUP_EXCEL_FILE = 'tfortis_backup.xlsx'
CAM_MAPPING_SHEET = 'cam'
CACHE_SHEET_SWITCH = 'switch_port_cache'
OUTPUT_FILE = 'result.xlsx'

CAMERA_NETWORK = ipaddress.IPv4Network('10.125.40.0/23')
PORT_RANGE = range(0, 6)  # 0–5

TEMP_IP_START = '10.125.41.220'
TEMP_IP_END = '10.125.41.230'

logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

def arp_scan(network):
    logger.info(f"Начинаю ARP-сканирование сети {network}...")
    try:
        ans, _ = srp(Ether(dst="ff:ff:ff:ff:ff:ff")/ARP(pdst=str(network)), timeout=2, verbose=0)
    except PermissionError:
        sys.exit(1)
    result = []
    for sent, received in ans:
        result.append({'ip': received.psrc, 'mac': received.hwsrc.upper().replace(':', '').replace('-', '')})
    logger.info(f"ARP-сканирование завершено, найдено {len(result)} устройств.")
    return result

def get_switch_mac(ip: str, port: int, attempts: int = 2, delay: float = 5.0) -> Optional[str]:
    """
    Пытается получить MAC с порта коммутатора до 'attempts' раз с паузой delay секунд.
    """
    url = f'http://{ip}/api/getPortMacList?port={port}'
    for attempt in range(attempts):
        try:
            resp = requests.get(url, auth=HTTPDigestAuth(SWITCH_USERNAME, SWITCH_PASSWORD), timeout=30)
            resp.raise_for_status()
            text = resp.text.strip()
            try:
                data = resp.json()
            except json.JSONDecodeError:
                start, end = text.find('{'), text.rfind('}')
                data = json.loads(text[start:end+1]) if start != -1 and end != -1 else {}
            if isinstance(data, dict) and '1' in data:
                v = data['1']
            else:
                lst = data.get(str(port), [])
                v = lst[1] if isinstance(lst, list) and len(lst) > 1 else None
            if isinstance(v, str) and v:
                return v.upper().replace(':', '').replace('-', '')
        except Exception as e:
            if attempt < attempts - 1:
                logger.warning(f"Попытка {attempt+1} неудачна для {ip}:{port}, пробую ещё раз...")
                time.sleep(delay)
            else:
                logger.error(f"Ошибка получения MAC {ip} порт {port} после {attempts} попыток: {e}")
    return None


def load_cache(sheet: str) -> pd.DataFrame:
    try:
        return pd.read_excel(EXCEL_FILE, sheet_name=sheet, dtype=str)
    except:
        return pd.DataFrame()

def now():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def collect_switch_ports(sw):
    """Опрашивает все порты одного коммутатора (последовательно), возвращает список записей"""
    recs = []
    for port in PORT_RANGE:
        mac = get_switch_mac(sw, port)
        if mac:
            recs.append({'Switch IP': sw, 'Port': str(port), 'MAC': mac, 'LastChecked': now()})
        time.sleep(10)
    return recs

def backup_excel():
    if os.path.exists(EXCEL_FILE):
        shutil.copy(EXCEL_FILE, BACKUP_EXCEL_FILE)
        logger.info(f"Сделана резервная копия файла: {BACKUP_EXCEL_FILE}")

def set_camera_ip(ip: str, new_ip: str, mask: str, gw: str) -> bool:
    url = f"http://{ip}/cgi-bin/configManager.cgi"
    params = {
        'action': 'setConfig',
        'Network.eth0.IPAddress': new_ip,
        'Network.eth0.SubnetMask': mask,
        'Network.eth0.DefaultGateway': gw,
    }
    try:
        resp = requests.get(url, params=params, auth=HTTPDigestAuth(CAMERA_USERNAME, CAMERA_PASSWORD), timeout=20)
        resp.raise_for_status()
        if "OK" in resp.text.upper() or "<response>OK</response>" in resp.text.upper():
            logger.info(f"IP камеры {ip} меняется на {new_ip} (успешно)")
            time.sleep(2)
            return True
        else:
            logger.warning(f"Камера {ip} не подтвердила смену IP! Ответ: {resp.text.strip()}")
            return False
    except Exception as e:
        logger.error(f"Ошибка смены IP на {ip}: {e}")
        return False

def parse_ip_range(ip_range_str):
    try:
        start_ip, end_ip = ip_range_str.replace(' ', '').split('-')
        start_ip, end_ip = ipaddress.IPv4Address(start_ip), ipaddress.IPv4Address(end_ip)
        return [str(ipaddress.IPv4Address(ip)) for ip in range(int(start_ip), int(end_ip)+1)]
    except Exception as e:
        logger.error(f"Ошибка разбора диапазона IP: {e}")
        return []

def find_free_ip(df_map, start=TEMP_IP_START, end=TEMP_IP_END):
    used_ips = set(df_map['Actual Camera IP'].dropna().tolist() + df_map['Expected Camera IP'].dropna().tolist())
    for ip_int in range(int(ipaddress.IPv4Address(start)), int(ipaddress.IPv4Address(end)) + 1):
        ip = str(ipaddress.IPv4Address(ip_int))
        if ip not in used_ips:
            return ip
    return None

def find_row_by_actual_ip(df_map, ip):
    for idx, row in df_map.iterrows():
        if row['Actual Camera IP'] == ip:
            return idx
    return None

def find_row_by_expected_ip(df_map, ip):
    for idx, row in df_map.iterrows():
        if row['Expected Camera IP'] == ip:
            return idx
    return None

def process_ip_move(df_map, from_idx, processed_idxs, mask, gw, allowed_ips):
    if from_idx in processed_idxs:
        row = df_map.iloc[from_idx]
        current_ip = row['Actual Camera IP']
        tmp_ip = find_free_ip(df_map)
        if not tmp_ip:
            logger.error(f"Нет свободного временного IP для разрыва цикла смены для {current_ip}")
            return False
        logger.warning(f"Обнаружен цикл! Временно переносим камеру {current_ip} на {tmp_ip} для разрыва перестановки.")
        ok = set_camera_ip(current_ip, tmp_ip, mask, gw)
        if ok:
            df_map.at[from_idx, 'Actual Camera IP'] = tmp_ip
            logger.info(f"Успешно временно сменён IP {current_ip} -> {tmp_ip}")
            return True
        else:
            logger.error(f"Ошибка при временной смене IP {current_ip} -> {tmp_ip}")
            return False
    processed_idxs.add(from_idx)
    row = df_map.iloc[from_idx]
    current_ip = row['Actual Camera IP']
    planned_ip = row['Expected Camera IP']
    if not planned_ip or not current_ip or current_ip == planned_ip:
        return True
    if allowed_ips and planned_ip not in allowed_ips:
        return True

    conflict_idx = find_row_by_actual_ip(df_map, planned_ip)
    if conflict_idx is not None:
        ok = process_ip_move(df_map, conflict_idx, processed_idxs, mask, gw, allowed_ips)
        if not ok:
            logger.error(f"Не удалось освободить IP {planned_ip} (занят строкой {conflict_idx+1}), {current_ip} пока не трогаем")
            return False

    ok = set_camera_ip(current_ip, planned_ip, mask, gw)
    if ok:
        logger.info(f"Успешно сменён IP {current_ip} -> {planned_ip}")
        df_map.at[from_idx, 'Actual Camera IP'] = planned_ip
        return True
    else:
        logger.error(f"Ошибка смены IP {current_ip} -> {planned_ip}")
        return False

def change_ip_according_to_plan(df_map):
    mask = input("Введите маску подсети (например, 255.255.254.0): ").strip()
    gw = input("Введите шлюз (например, 10.125.41.254): ").strip()
    ip_range = input("Введите диапазон ожидаемых IP (например, 10.125.40.1-10.125.40.50): ").strip()
    if ip_range:
        allowed_ips = set(parse_ip_range(ip_range))
        logger.info(f"Будут обработаны только камеры с ожидаемым IP из диапазона: {ip_range}")
    else:
        allowed_ips = None

    planned_ip_list = [row['Expected Camera IP'] for _, row in df_map.iterrows() if row['Expected Camera IP']]
    duplicated_ips = set([ip for ip in planned_ip_list if planned_ip_list.count(ip) > 1])
    if duplicated_ips:
        logger.error(f"ВНИМАНИЕ! Повторяющиеся ожидаемые IP: {', '.join(duplicated_ips)}")
        print(f"ОШИБКА: В таблице есть повторяющиеся IP. Исправьте их перед массовой сменой!")
        return

    changed = 0
    for idx, row in df_map.iterrows():
        current_ip = row['Actual Camera IP']
        planned_ip = row['Expected Camera IP']
        if not planned_ip or not current_ip or current_ip == planned_ip:
            continue
        if allowed_ips and planned_ip not in allowed_ips:
            continue
        processed_idxs = set()
        ok = process_ip_move(df_map, idx, processed_idxs, mask, gw, allowed_ips)
        if ok:
            changed += 1
    logger.info(f"Всего сменено адресов: {changed}")

def style_result_xlsx(path):
    wb = load_workbook(path)
    ws = wb['mapping']
    status_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Camera Status':
            status_col = col_idx
    if status_col:
        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        for row in ws.iter_rows(min_row=2):
            status = (row[status_col-1].value or '').lower()
            if status == 'online':
                for cell in row: cell.fill = green
            elif status == 'offline':
                for cell in row: cell.fill = red
            elif status == 'unknown':
                for cell in row: cell.fill = yellow
    wb.save(path)

def main():
    update_switch_cache = input("Обновить кэш портов коммутаторов? (y/N): ").strip().lower() == 'y'
    if update_switch_cache:
        backup_excel()

    df_map = pd.read_excel(EXCEL_FILE, sheet_name=CAM_MAPPING_SHEET, dtype=str)
    for col in ['Пикет','Switch name','Switch IP','Установка','System name']:
        if col in df_map.columns:
            df_map[col] = df_map[col].ffill()
    df_map.rename(columns={'IP':'Expected Camera IP'}, inplace=True)
    df_map['Actual Camera IP'] = ''
    df_map['Camera Status'] = ''
    df_map['Port'] = ''
    df_map['Port MAC'] = ''
    df_map['LastChecked'] = ''
    df_map['Offline Last IP'] = ''
    df_map['Offline Last Seen'] = ''

    # === ARP Сканирование ===
    arp_list = arp_scan(CAMERA_NETWORK)
    arp_ip_to_mac = {dev['ip']: dev['mac'] for dev in arp_list}
    arp_mac_to_ip = {dev['mac']: dev['ip'] for dev in arp_list}

    # === КЭШ ПОРТОВ КОММУТАТОРОВ ===
    if update_switch_cache:
        logger.info("Обновляем кэш портов коммутаторов (многопоточно)...")
        switches = df_map['Switch IP'].dropna().unique().tolist()
        port_records = []
        with ThreadPoolExecutor(max_workers=50) as pool:
            futures = {pool.submit(collect_switch_ports, sw): sw for sw in switches}
            for fut in concurrent.futures.as_completed(futures):
                sw = futures[fut]
                try:
                    recs = fut.result()
                    port_records.extend(recs)
                except Exception as e:
                    logger.warning(f"Ошибка при опросе коммутатора {sw}: {e}")
        df_spc = pd.DataFrame(port_records)
        old_spc = load_cache(CACHE_SHEET_SWITCH)
        if not old_spc.empty:
            df_spc = pd.concat([old_spc, df_spc]).drop_duplicates(['Switch IP', 'Port'], keep='last').reset_index(drop=True)
        with pd.ExcelWriter(EXCEL_FILE, mode='a', if_sheet_exists='replace') as writer:
            df_spc.to_excel(writer, sheet_name=CACHE_SHEET_SWITCH, index=False)
        logger.info("Кэш портов обновлен: %d строк", len(df_spc))
    else:
        df_spc = load_cache(CACHE_SHEET_SWITCH)



    # === Заполнение итоговой таблицы ===
    unknown_devices = []
    for idx, row in df_map.iterrows():
        sw = row['Switch IP']
        cameras_for_sw = df_map[df_map['Switch IP'] == sw].reset_index()
        port = cameras_for_sw[cameras_for_sw['index'] == idx].index[0] if sw in df_spc['Switch IP'].values else None
        df_map.at[idx, 'Port'] = port if port is not None else ''
        port_mac = ''
        port_lastchecked = ''
        actual_ip = ''
        cam_status = 'offline'
        offline_last_ip = ''
        offline_last_seen = ''

        if port is not None:
            port_row = df_spc[(df_spc['Switch IP'] == sw) & (df_spc['Port'] == str(port))]
            if not port_row.empty:
                port_mac = port_row.iloc[0]['MAC']
                port_lastchecked = port_row.iloc[0]['LastChecked']
                # Вся логика поиска IP только через ARP
                actual_ip = arp_mac_to_ip.get(port_mac, '')
                if actual_ip:
                    cam_status = 'online'
                else:
                    unknown_devices.append({'Switch IP': sw, 'Port': port, 'MAC': port_mac, 'LastChecked': port_lastchecked})
        df_map.at[idx, 'Port MAC'] = port_mac
        df_map.at[idx, 'Actual Camera IP'] = actual_ip
        df_map.at[idx, 'Camera Status'] = cam_status if actual_ip else ("unknown" if port_mac else "offline")
        df_map.at[idx, 'Offline Last IP'] = offline_last_ip
        df_map.at[idx, 'Offline Last Seen'] = offline_last_seen
        if not port_lastchecked and port_mac:
            df_map.at[idx, 'LastChecked'] = port_lastchecked

    df_unknown = pd.DataFrame(unknown_devices)

    if input('Сделать массовую смену IP по плану? (y/N): ').strip().lower() == 'y':
        change_ip_according_to_plan(df_map)

    summary = df_map.groupby('Switch IP')['Actual Camera IP'] \
                    .apply(lambda x: ', '.join(filter(None, x))) \
                    .reset_index() \
                    .rename(columns={'Actual Camera IP': 'Camera IPs'})
    with pd.ExcelWriter(OUTPUT_FILE) as writer:
        df_map.to_excel(writer, sheet_name='mapping', index=False)
        summary.to_excel(writer, sheet_name='summary', index=False)
        df_spc.to_excel(writer, sheet_name='switch_port_cache', index=False)
        df_unknown.to_excel(writer, sheet_name='unknown_devices', index=False)
    logger.info("Результат сохранен в %s", OUTPUT_FILE)

    style_result_xlsx(OUTPUT_FILE)

if __name__ == '__main__':
    main()
